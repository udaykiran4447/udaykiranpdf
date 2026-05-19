import streamlit as st

st.set_page_config(
    page_title="Compliance Suite",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ─── Session state ──────────────────────────────────────────────────────────────
if "active_app" not in st.session_state:
    st.session_state.active_app = None
if "theme" not in st.session_state:
    st.session_state.theme = "dark"

is_dark = st.session_state.theme == "dark"

# ── Theme toggle (top-right) ────────────────────────────────────────────────────
_tc1, _tc2 = st.columns([10, 1])
with _tc2:
    if st.button("☀️" if is_dark else "🌙", key="theme_toggle", help="Toggle light/dark mode"):
        st.session_state.theme = "light" if is_dark else "dark"
        st.rerun()

# ── Design tokens ───────────────────────────────────────────────────────────────
if is_dark:
    T = dict(
        app_bg="#0e0e16",
        hero_bg="linear-gradient(135deg,#0d0d1a 0%,#131330 60%,#0a0f1e 100%)",
        hero_border="rgba(255,255,255,0.08)",
        card_bg="#16161f",
        card_border="rgba(255,255,255,0.10)",
        card_border_h="rgba(255,255,255,0.26)",
        card_shadow="rgba(0,0,0,0.55)",
        title_col="#eeeeff",
        sub_col="#b8bcce",
        feat_col="#c4c8dc",
        format_col="#9095ad",
        section_label="#7a7f97",
        footer_border="rgba(255,255,255,0.08)",
        badge_bg="rgba(255,255,255,0.08)",
        badge_border="rgba(255,255,255,0.15)",
        badge_col="#c4c8dc",
        badge_b="#eeeeff",
        eyebrow_bg="rgba(99,102,241,0.18)",
        eyebrow_brd="rgba(99,102,241,0.45)",
        eyebrow_col="#c7caff",
        glow1="rgba(99,102,241,0.22)",
        glow2="rgba(16,185,129,0.14)",
        nav_bg="#0d0d1a",
        nav_border="rgba(255,255,255,0.08)",
        nav_title="#eeeeff",
        hr="rgba(255,255,255,0.08)",
        toggle_bg="#1c1c2c",
        toggle_brd="rgba(255,255,255,0.14)",
        toggle_col="#c4c8dc",
        grid_bg="#0e0e16",
        ind_tag_bg="rgba(99,102,241,0.18)",  ind_tag_col="#b0b3ff",  ind_icon_bg="rgba(99,102,241,0.18)",
        em_tag_bg="rgba(16,185,129,0.15)",   em_tag_col="#6ee7b7",   em_icon_bg="rgba(16,185,129,0.15)",
        am_tag_bg="rgba(245,158,11,0.18)",   am_tag_col="#fcd34d",   am_icon_bg="rgba(245,158,11,0.15)",
        ro_tag_bg="rgba(244,63,94,0.16)",    ro_tag_col="#fda4af",   ro_icon_bg="rgba(244,63,94,0.15)",
    )
else:
    T = dict(
        app_bg="#f0f2f7",
        hero_bg="linear-gradient(135deg,#1a1a3e 0%,#1e2a6e 60%,#0f1f5c 100%)",
        hero_border="rgba(255,255,255,0.15)",
        card_bg="#ffffff",
        card_border="rgba(0,0,0,0.07)",
        card_border_h="rgba(0,0,0,0.20)",
        card_shadow="rgba(0,0,0,0.10)",
        title_col="#0f172a",
        sub_col="#334155",
        feat_col="#374151",
        format_col="#64748b",
        section_label="#64748b",
        footer_border="rgba(0,0,0,0.08)",
        badge_bg="rgba(255,255,255,0.18)",
        badge_border="rgba(255,255,255,0.32)",
        badge_col="#e2e8f0",
        badge_b="#ffffff",
        eyebrow_bg="rgba(255,255,255,0.18)",
        eyebrow_brd="rgba(255,255,255,0.38)",
        eyebrow_col="#dde3ff",
        glow1="rgba(99,102,241,0.28)",
        glow2="rgba(16,185,129,0.18)",
        nav_bg="#ffffff",
        nav_border="rgba(0,0,0,0.07)",
        nav_title="#0f172a",
        hr="rgba(0,0,0,0.08)",
        toggle_bg="#ffffff",
        toggle_brd="rgba(0,0,0,0.12)",
        toggle_col="#334155",
        grid_bg="#f0f2f7",
        ind_tag_bg="rgba(99,102,241,0.10)",  ind_tag_col="#3730a3",  ind_icon_bg="rgba(99,102,241,0.10)",
        em_tag_bg="rgba(16,185,129,0.10)",   em_tag_col="#065f46",   em_icon_bg="rgba(16,185,129,0.10)",
        am_tag_bg="rgba(245,158,11,0.12)",   am_tag_col="#78350f",   am_icon_bg="rgba(245,158,11,0.10)",
        ro_tag_bg="rgba(244,63,94,0.10)",    ro_tag_col="#9f1239",   ro_icon_bg="rgba(244,63,94,0.10)",
    )

hero_sub_color = "#c0c4d8" if is_dark else "#334155"

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Sans:wght@300;400;500;600&display=swap');
html,body,[class*="css"]{{font-family:'DM Sans',sans-serif;}}
.stApp{{background:{T['app_bg']};min-height:100vh;}}
#MainMenu,footer,header{{visibility:hidden;}}
.block-container{{padding:0!important;max-width:100%!important;}}

/* toggle btn */
section[data-testid="stSidebar"]{{display:none;}}
div[data-testid="column"]:last-child .stButton>button{{
    background:{T['toggle_bg']}!important;border:1px solid {T['toggle_brd']}!important;
    border-radius:100px!important;color:{T['toggle_col']}!important;
    font-size:1.1rem!important;font-weight:600!important;padding:4px 14px!important;
    box-shadow:0 2px 10px rgba(0,0,0,.15)!important;margin-top:8px!important;
    width:auto!important;float:right!important;
}}

/* hero */
.hero-wrap{{position:relative;overflow:hidden;background:{T['hero_bg']};
    padding:3.2rem 3rem 2.6rem;border-bottom:1px solid {T['hero_border']};}}
.hero-wrap::before{{content:'';position:absolute;top:-80px;right:-100px;
    width:480px;height:480px;border-radius:50%;pointer-events:none;
    background:radial-gradient(circle,{T['glow1']} 0%,transparent 70%);}}
.hero-wrap::after{{content:'';position:absolute;bottom:-60px;left:40px;
    width:300px;height:300px;border-radius:50%;pointer-events:none;
    background:radial-gradient(circle,{T['glow2']} 0%,transparent 70%);}}
.hero-eyebrow{{display:inline-flex;align-items:center;gap:8px;
    background:{T['eyebrow_bg']};border:1px solid {T['eyebrow_brd']};
    border-radius:100px;padding:5px 16px;font-size:.72rem;font-weight:700;
    letter-spacing:.13em;text-transform:uppercase;color:{T['eyebrow_col']};margin-bottom:1.2rem;}}
.hero-title{{font-family:'Syne',sans-serif;font-size:3rem;font-weight:800;
    line-height:1.1;color:#f0f0ff;margin:0 0 .9rem;letter-spacing:-.02em;}}
.hero-title span{{background:linear-gradient(90deg,#818cf8,#34d399);
    -webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;}}
.hero-sub{{color:{hero_sub_color};font-size:1.05rem;font-weight:400;max-width:560px;line-height:1.65;margin:0;}}
.hero-badges{{display:flex;gap:10px;margin-top:1.6rem;flex-wrap:wrap;}}
.hero-badge{{background:{T['badge_bg']};border:1px solid {T['badge_border']};
    border-radius:7px;padding:6px 14px;font-size:.78rem;color:{T['badge_col']};font-weight:500;}}
.hero-badge b{{color:{T['badge_b']};}}

/* grid */
.grid-section{{padding:2.5rem 3rem 3.5rem;background:{T['grid_bg']};}}
.section-label{{font-family:'Syne',sans-serif;font-size:.72rem;font-weight:700;
    letter-spacing:.14em;text-transform:uppercase;color:{T['section_label']};margin-bottom:1.5rem;}}

/* card */
.app-card{{background:{T['card_bg']};border:1px solid {T['card_border']};
    border-radius:18px;overflow:hidden;
    transition:transform .22s ease,border-color .22s ease,box-shadow .22s ease;}}
.app-card:hover{{transform:translateY(-5px);border-color:{T['card_border_h']};
    box-shadow:0 22px 55px {T['card_shadow']};}}
.card-accent{{height:4px;width:100%;}}
.card-body{{padding:1.7rem 1.9rem 1.5rem;}}
.card-icon-row{{display:flex;align-items:center;gap:1rem;margin-bottom:1.1rem;}}
.card-icon{{width:50px;height:50px;border-radius:13px;display:flex;align-items:center;justify-content:center;font-size:1.45rem;flex-shrink:0;}}
.card-tag{{font-size:.7rem;font-weight:700;letter-spacing:.1em;text-transform:uppercase;padding:4px 11px;border-radius:5px;}}
.card-title{{font-family:'Syne',sans-serif;font-size:1.28rem;font-weight:700;color:{T['title_col']};margin:0 0 .55rem;letter-spacing:-.01em;}}
.card-desc{{font-size:.89rem;color:{T['sub_col']};line-height:1.65;margin:0 0 1.2rem;font-weight:400;}}
.card-features{{list-style:none;padding:0;margin:0 0 1.3rem;display:flex;flex-direction:column;gap:7px;}}
.card-features li{{font-size:.83rem;color:{T['feat_col']};display:flex;align-items:center;gap:9px;font-weight:400;}}
.card-features li::before{{content:'';width:6px;height:6px;border-radius:50%;flex-shrink:0;}}
.card-footer{{display:flex;align-items:center;justify-content:space-between;
    padding-top:1rem;border-top:1px solid {T['footer_border']};}}
.card-format{{font-size:.76rem;color:{T['format_col']};font-weight:500;}}

/* themes */
.theme-indigo .card-accent{{background:linear-gradient(90deg,#6366f1,#818cf8);}}
.theme-indigo .card-icon{{background:{T['ind_icon_bg']};}}
.theme-indigo .card-tag{{background:{T['ind_tag_bg']};color:{T['ind_tag_col']};}}
.theme-indigo .card-features li::before{{background:#6366f1;}}

.theme-emerald .card-accent{{background:linear-gradient(90deg,#10b981,#34d399);}}
.theme-emerald .card-icon{{background:{T['em_icon_bg']};}}
.theme-emerald .card-tag{{background:{T['em_tag_bg']};color:{T['em_tag_col']};}}
.theme-emerald .card-features li::before{{background:#10b981;}}

.theme-amber .card-accent{{background:linear-gradient(90deg,#f59e0b,#fbbf24);}}
.theme-amber .card-icon{{background:{T['am_icon_bg']};}}
.theme-amber .card-tag{{background:{T['am_tag_bg']};color:{T['am_tag_col']};}}
.theme-amber .card-features li::before{{background:#f59e0b;}}

.theme-rose .card-accent{{background:linear-gradient(90deg,#f43f5e,#fb7185);}}
.theme-rose .card-icon{{background:{T['ro_icon_bg']};}}
.theme-rose .card-tag{{background:{T['ro_tag_bg']};color:{T['ro_tag_col']};}}
.theme-rose .card-features li::before{{background:#f43f5e;}}

/* launch btns */
.stButton>button{{font-family:'DM Sans',sans-serif!important;font-weight:600!important;
    font-size:.86rem!important;border-radius:0 0 16px 16px!important;border:none!important;
    padding:.7rem 1rem!important;width:100%!important;
    transition:opacity .15s,transform .15s!important;letter-spacing:.01em!important;}}
.stButton>button:hover{{opacity:.87!important;transform:translateY(-1px)!important;}}

/* nav */
.nav-title{{font-family:'Syne',sans-serif;font-size:1rem;font-weight:700;color:{T['nav_title']};padding-top:4px;}}
</style>
""", unsafe_allow_html=True)

# ── App list ─────────────────────────────────────────────────────────────────
APPS = [
    dict(id="tds",  theme="indigo",  icon="📄", tag="TDS",
         title="TDS Challan Extractor",
         desc="Parse ITNS 281 challan PDFs and export structured data to a formatted Excel workbook.",
         features=["Extracts TAN, CIN, BSR Code, Challan No","Tax, Surcharge, Cess & Fee breakdowns","Multi-file batch processing","Auto-totals Excel export"],
         fmt="PDF → Excel"),
    dict(id="epf",  theme="emerald", icon="📋", tag="EPF",
         title="EPF Challan Consolidator",
         desc="Consolidate multiple EPFO Combined Challan PDFs into one detailed report with A/C breakdowns.",
         features=["A/C 01, 02, 10, 21 & 22 parsing","Employer & Employee share split","Grand Total + wages summary","Per-establishment expanders"],
         fmt="PDF → Excel"),
    dict(id="esic", theme="amber",   icon="🏥", tag="ESIC",
         title="ESIC Challan Extractor",
         desc="Extract employer code, challan number, period and payment info from ESIC challan PDFs.",
         features=["Employer code & name extraction","Challan period & transaction details","Amount Paid with totals row","Clean formatted Excel output"],
         fmt="PDF → Excel"),
    dict(id="excel",theme="rose",    icon="📊", tag="Excel",
         title="Excel File Consolidator",
         desc="Merge XLS, XLSX, XLSM and CSV files of any format into one clean spreadsheet with source tracking.",
         features=["Handles binary XLS & HTML-as-XLS","Union or intersection column strategy","Source File column auto-added","File Summary sheet included"],
         fmt="XLS / XLSX / CSV → Excel"),
]
BTN_BG = {"indigo":"#6366f1","emerald":"#10b981","amber":"#f59e0b","rose":"#f43f5e"}
BTN_FG = {"indigo":"#fff","emerald":"#fff","amber":"#0a0a0f","rose":"#fff"}

# ══════════════════════════════════════════════════════════════════════════════
# HOME
# ══════════════════════════════════════════════════════════════════════════════
if st.session_state.active_app is None:
    st.markdown(f"""
    <div class="hero-wrap">
        <div class="hero-eyebrow">⚡ Compliance Automation Suite</div>
        <h1 class="hero-title">Your payroll compliance<br><span>tools, unified.</span></h1>
        <p class="hero-sub">Extract, consolidate and export TDS, EPF, ESIC challans and Excel files — all in one place.</p>
        <div class="hero-badges">
            <span class="hero-badge"><b>4</b> tools</span>
            <span class="hero-badge"><b>PDF</b> extraction</span>
            <span class="hero-badge"><b>Excel</b> export</span>
            <span class="hero-badge"><b>Batch</b> processing</span>
        </div>
    </div>
    <div class="grid-section">
        <div class="section-label">Select a tool to get started</div>
    """, unsafe_allow_html=True)

    col_a, col_b = st.columns(2, gap="large")
    cols = [col_a, col_b]
    for i, app in enumerate(APPS):
        feats = "".join(f"<li>{f}</li>" for f in app["features"])
        with cols[i % 2]:
            st.markdown(f"""
            <div class="app-card theme-{app['theme']}">
                <div class="card-accent"></div>
                <div class="card-body">
                    <div class="card-icon-row">
                        <div class="card-icon">{app['icon']}</div>
                        <span class="card-tag">{app['tag']}</span>
                    </div>
                    <div class="card-title">{app['title']}</div>
                    <div class="card-desc">{app['desc']}</div>
                    <ul class="card-features">{feats}</ul>
                    <div class="card-footer">
                        <span class="card-format">{app['fmt']}</span>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)
            bg = BTN_BG[app['theme']]; fg = BTN_FG[app['theme']]
            st.markdown(f"<style>div[data-testid='stVerticalBlock'] div[data-testid='stVerticalBlock']:nth-child({i+2}) .stButton>button{{background:{bg}!important;color:{fg}!important;}}</style>", unsafe_allow_html=True)
            if st.button(f"Launch {app['title']} →", key=f"launch_{app['id']}", use_container_width=True):
                st.session_state.active_app = app["id"]
                st.rerun()

    st.markdown("</div>", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# ACTIVE APP
# ══════════════════════════════════════════════════════════════════════════════
else:
    active = next(a for a in APPS if a["id"] == st.session_state.active_app)
    nb1, nb2, _ = st.columns([1.5, 6, 3])
    with nb1:
        if st.button("← Back", key="nav_back"):
            st.session_state.active_app = None
            for k in list(st.session_state.keys()):
                if k not in ("active_app","theme"):
                    del st.session_state[k]
            st.rerun()
    with nb2:
        st.markdown(f"<div class='nav-title'>{active['icon']}  {active['title']}</div>", unsafe_allow_html=True)
    st.markdown(f"<hr style='border:none;border-top:1px solid {T['hr']};margin:.2rem 0 1.2rem'>", unsafe_allow_html=True)

    # ═══════════════════════════════════════════════════════════════════════════
    # TDS
    # ═══════════════════════════════════════════════════════════════════════════
    if active["id"] == "tds":
        import pdfplumber, pandas as pd, re
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        st.markdown(f"""<style>
        .main-header{{font-size:1.7rem;font-weight:700;color:{T['title_col']};margin-bottom:4px;font-family:'Syne',sans-serif;}}
        .sub-header{{font-size:.9rem;color:{T['sub_col']};margin-bottom:1.2rem;}}
        .metric-card{{background:{T['card_bg']};border-radius:10px;padding:1rem 1.2rem;text-align:center;border:1px solid {T['card_border']};}}
        .metric-value{{font-size:1.8rem;font-weight:700;color:{T['title_col']};}}
        .metric-label{{font-size:.75rem;color:{T['sub_col']};margin-top:4px;font-weight:500;}}
        </style>""", unsafe_allow_html=True)

        st.markdown('<div class="main-header">📄 TDS Challan PDF Extractor</div>', unsafe_allow_html=True)
        st.markdown('<div class="sub-header">Upload ITNS 281 challan receipts — all data extracted into a single Excel sheet</div>', unsafe_allow_html=True)

        def tds_ev(text, label):
            for p in [rf"{re.escape(label)}\s*[:\-]\s*(.+)", rf"{re.escape(label)}\s+(.+)"]:
                m = re.search(p, text, re.IGNORECASE)
                if m: return m.group(1).strip()
            return ""

        def tds_ca(val):
            val = val.replace("₹","").replace(",","").strip()
            m = re.search(r"[\d]+(?:\.\d+)?", val)
            return float(m.group()) if m else 0.0

        def tds_extract(f):
            with pdfplumber.open(f) as pdf:
                txt = "".join((p.extract_text() or "")+"\n" for p in pdf.pages)
            d = {}
            for lbl,key in {"TAN":"TAN","Name":"Name","Assessment Year":"Assessment Year",
                            "Financial Year":"Financial Year","Major Head":"Major Head","Minor Head":"Minor Head",
                            "Nature of Payment":"Nature of Payment","CIN":"CIN","Mode of Payment":"Mode of Payment",
                            "Bank Name":"Bank Name","Bank Reference Number":"Bank Reference Number",
                            "Date of Deposit":"Date of Deposit","BSR code":"BSR Code",
                            "Challan No":"Challan No","Tender Date":"Tender Date"}.items():
                d[key] = tds_ev(txt, lbl)
            am = re.search(r"Amount \(in Rs\.\)\s*[:\-]?\s*₹?\s*([\d,]+)", txt)
            d["Amount (Rs.)"] = tds_ca(am.group(1)) if am else 0.0
            aw = re.search(r"Amount \(in words\)\s*[:\-]?\s*(.+)", txt)
            d["Amount (in words)"] = aw.group(1).strip() if aw else ""
            for k,p in {"Tax":r"A\s+Tax\s+₹?\s*([\d,]+)","Surcharge":r"B\s+Surcharge\s+₹?\s*([\d,]+)",
                        "Cess":r"C\s+Cess\s+₹?\s*([\d,]+)","Interest":r"D\s+Interest\s+₹?\s*([\d,]+)",
                        "Penalty":r"E\s+Penalty\s+₹?\s*([\d,]+)",
                        "Fee u/s 234E":r"F\s+Fee under section 234E\s+₹?\s*([\d,]+)",
                        "Total":r"Total \(A\+B\+C\+D\+E\+F\)\s+₹?\s*([\d,]+)"}.items():
                m = re.search(p, txt); d[k] = tds_ca(m.group(1)) if m else 0.0
            im = re.search(r"ITNS No\.\s*[:\-]?\s*(\d+)", txt)
            d["ITNS No."] = im.group(1).strip() if im else ""
            return d

        def tds_excel(records):
            wb=Workbook(); ws=wb.active; ws.title="TDS Challans"
            HF=PatternFill("solid",start_color="1a1a2e",end_color="1a1a2e")
            HFn=Font(bold=True,color="FFFFFF",name="Arial",size=10)
            SF=PatternFill("solid",start_color="E8F4FD",end_color="E8F4FD")
            SFn=Font(bold=True,name="Arial",size=9,color="1a1a2e")
            DF=Font(name="Arial",size=9)
            AF=PatternFill("solid",start_color="F8F9FA",end_color="F8F9FA")
            C=Alignment(horizontal="center",vertical="center")
            L=Alignment(horizontal="left",vertical="center")
            thin=Side(style="thin",color="DEE2E6")
            B=Border(left=thin,right=thin,top=thin,bottom=thin)
            ws.merge_cells("A1:T1"); ws["A1"]="TDS CHALLAN DETAILS — KAPSTON SERVICES LIMITED"
            ws["A1"].font=Font(bold=True,name="Arial",size=12,color="FFFFFF"); ws["A1"].fill=HF; ws["A1"].alignment=C; ws.row_dimensions[1].height=28
            hdrs=["S.No","ITNS No.","TAN","Name","Assessment Year","Financial Year","Nature of Payment","CIN",
                  "Mode of Payment","Bank Name","Bank Ref. No.","Date of Deposit","BSR Code","Challan No","Tender Date",
                  "Tax (Rs.)","Surcharge (Rs.)","Cess (Rs.)","Interest (Rs.)","Penalty (Rs.)","Fee u/s 234E (Rs.)","Total Amount (Rs.)"]
            ws.merge_cells("A2:A3"); ws.merge_cells("B2:B3")
            for ci,h in enumerate(hdrs,1):
                c=ws.cell(row=2,column=ci,value=h); c.font=HFn; c.fill=HF; c.alignment=C; c.border=B
                ws.cell(row=3,column=ci).font=SFn; ws.cell(row=3,column=ci).fill=SF
                ws.cell(row=3,column=ci).alignment=C; ws.cell(row=3,column=ci).border=B
            ws.row_dimensions[2].height=20; ws.row_dimensions[3].height=16
            for i,rec in enumerate(records):
                row=i+4; fill=AF if i%2==0 else PatternFill("solid",start_color="FFFFFF",end_color="FFFFFF")
                vals=[i+1,rec.get("ITNS No.",""),rec.get("TAN",""),rec.get("Name",""),
                      rec.get("Assessment Year",""),rec.get("Financial Year",""),rec.get("Nature of Payment",""),
                      rec.get("CIN",""),rec.get("Mode of Payment",""),rec.get("Bank Name",""),
                      rec.get("Bank Reference Number",""),rec.get("Date of Deposit",""),rec.get("BSR Code",""),
                      rec.get("Challan No",""),rec.get("Tender Date",""),rec.get("Tax",0),rec.get("Surcharge",0),
                      rec.get("Cess",0),rec.get("Interest",0),rec.get("Penalty",0),rec.get("Fee u/s 234E",0),rec.get("Total",0)]
                for ci,v in enumerate(vals,1):
                    c=ws.cell(row=row,column=ci,value=v); c.font=DF; c.fill=fill; c.border=B
                    c.alignment=C if ci==1 else L
                    if ci>=16: c.number_format='₹#,##0.00'
                ws.row_dimensions[row].height=18
            tr=len(records)+4
            ws.cell(row=tr,column=1,value="TOTAL").fill=HF
            ws.cell(row=tr,column=1).font=Font(bold=True,color="FFFFFF",name="Arial",size=9)
            ws.cell(row=tr,column=1).alignment=C; ws.merge_cells(f"A{tr}:O{tr}")
            for ci in range(16,23):
                cl=get_column_letter(ci); c=ws.cell(row=tr,column=ci,value=f"=SUM({cl}4:{cl}{tr-1})")
                c.font=Font(bold=True,name="Arial",size=9,color="FFFFFF"); c.fill=HF
                c.number_format='₹#,##0.00'; c.alignment=C; c.border=B
            ws.row_dimensions[tr].height=20
            for i,w in enumerate([5,8,14,28,14,12,18,26,14,14,18,14,10,10,12,14,14,10,10,10,14,16],1):
                ws.column_dimensions[get_column_letter(i)].width=w
            ws.freeze_panes="A4"
            out=BytesIO(); wb.save(out); out.seek(0); return out

        uf=st.file_uploader("Upload challan PDF files",type=["pdf"],accept_multiple_files=True,key="tds_uploader")
        if uf:
            st.markdown("---")
            c1,c2,c3=st.columns(3)
            c1.markdown(f'<div class="metric-card"><div class="metric-value">{len(uf)}</div><div class="metric-label">Files Uploaded</div></div>',unsafe_allow_html=True)
            recs=[]; errs=[]
            with st.spinner("Extracting data..."):
                for f in uf:
                    try: d=tds_extract(f); d["_filename"]=f.name; recs.append(d)
                    except Exception as e: errs.append((f.name,str(e)))
            for fn,err in errs: st.error(f"❌ {fn}: {err}")
            if recs:
                tot=sum(r.get("Total",0) for r in recs)
                c2.markdown(f'<div class="metric-card"><div class="metric-value">{len(recs)}</div><div class="metric-label">Extracted Successfully</div></div>',unsafe_allow_html=True)
                c3.markdown(f'<div class="metric-card"><div class="metric-value">₹{tot:,.0f}</div><div class="metric-label">Total TDS Amount</div></div>',unsafe_allow_html=True)
                st.markdown("### 📊 Data Preview")
                df=pd.DataFrame(recs); df.insert(0,"S.No",range(1,len(df)+1)); df["File"]=df["_filename"]
                pc=["Nature of Payment","CIN","Challan No","Date of Deposit","BSR Code","Tax","Surcharge","Cess","Interest","Penalty","Fee u/s 234E","Total"]
                st.dataframe(df[["S.No","File"]+[c for c in pc if c in df.columns]],use_container_width=True,hide_index=True)
                st.download_button("⬇️ Download Excel",data=tds_excel(recs),file_name="TDS_Challans.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.info("👆 Upload one or more TDS challan PDF files to get started.")
            st.markdown("**Supported format:** ITNS 281 Challan Receipts from Income Tax Department")

    # ═══════════════════════════════════════════════════════════════════════════
    # EPF
    # ═══════════════════════════════════════════════════════════════════════════
    elif active["id"] == "epf":
        import pdfplumber, re, io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        def epf_extract(f):
            try:
                with pdfplumber.open(f) as pdf:
                    text="\n".join(p.extract_text() or "" for p in pdf.pages)
            except Exception as e: return None,f"Could not read PDF: {e}"
            if "EMPLOYEES' PROVIDENT FUND" not in text: return None,"Not an EPF challan"
            data={}
            def find(pat,flags=0):
                m=re.search(pat,text,flags); return m.group(1).strip() if m else ""
            data["TRRN"]=find(r"TRRN[:\s]*(\d+)"); data["ECR Id"]=find(r"ECR\s*Id\s*(\d+)"); data["LIN"]=find(r"LIN\s*[:\s]*(\d+)")
            m=re.search(r"Establishment Code\s*&\s*([A-Z0-9]+)\s+(.+?)\s+Dues for the wage month\s+(\w+)\s+(\d{4})",text)
            if m: data["Establishment Code"]=m.group(1); data["Company Name"]=m.group(2).strip(); data["Wage Month"]=f"{m.group(3)} {m.group(4)}"
            else: data["Establishment Code"]=data["Company Name"]=data["Wage Month"]=""
            m=re.search(r"Address\s*:\s*(.+?)(?=\nEPF|\nTotal)",text,re.DOTALL)
            data["Address"]=re.sub(r'\s+',' ',m.group(1)).strip() if m else ""
            def pip(pat):
                m=re.search(pat,text)
                return (int(m.group(1).replace(",","")),int(m.group(2).replace(",",""))) if m else ("","")
            a,b=pip(r"Total Subscribers\s*:\s*([\d,]+)\s+([\d,]+)"); data["Total Subscribers EPF"]=a; data["Total Subscribers EPS"]=b
            a,b=pip(r"Total Wages\s*:\s*([\d,]+)\s+([\d,]+)"); data["Total Wages EPF"]=a; data["Total Wages EPS"]=b
            def pr(lbl):
                m=re.search(lbl+r"\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)",text)
                return [int(g.replace(",","")) for g in m.groups()] if m else [""]*6
            for pfx,lbl in [("Admin",r"Administration Charges"),("Employer",r"Employer.s Share Of"),("Employee",r"Employee.s Share Of")]:
                for col,v in zip(["A/C.01","A/C.02","A/C.10","A/C.21","A/C.22","Total"],pr(lbl)): data[f"{pfx} {col}"]=v
            m=re.search(r"Grand Total\s*:\s*(.+?)\s+([\d,]+)\s*$",text,re.MULTILINE)
            if m: data["Grand Total (Words)"]=m.group(1).strip(); data["Grand Total"]=int(m.group(2).replace(",",""))
            else: data["Grand Total (Words)"]=data["Grand Total"]=""
            m=re.search(r"Total remittance by Employer.*?([\d,]+)\s*$",text,re.MULTILINE)
            data["Total Remittance by Employer"]=int(m.group(1).replace(",","")) if m else ""
            m=re.search(r"Total amount of uploaded ECR.*?([\d,]+)\s*$",text,re.MULTILINE)
            data["Total ECR Amount"]=int(m.group(1).replace(",","")) if m else ""
            return data,None

        def epf_excel(records):
            wb=Workbook(); ws=wb.active; ws.title="EPF Challan Summary"
            thin=Side(style="thin",color="BFBFBF"); bdr=Border(left=thin,right=thin,top=thin,bottom=thin)
            ctr=Alignment(horizontal="center",vertical="center",wrap_text=True)
            lft=Alignment(horizontal="left",vertical="center",wrap_text=True)
            rgt=Alignment(horizontal="right",vertical="center")
            def mc(row,col,value=None,bold=False,fg="000000",bg=None,align=None,fmt=None,size=9):
                c=ws.cell(row=row,column=col,value=value); c.font=Font(name="Arial",bold=bold,color=fg,size=size)
                if bg: c.fill=PatternFill("solid",start_color=bg)
                c.alignment=align or lft; c.border=bdr
                if fmt: c.number_format=fmt
                return c
            INR='#,##0;(#,##0);"-"'; INT="#,##0"
            COLS=[("S.No",5,"S.No",False,True),("Source File",30,"Source File",False,False),
                  ("Establishment\nCode",20,"Establishment Code",False,False),("Company Name",28,"Company Name",False,False),
                  ("Address",42,"Address",False,False),("Wage\nMonth",13,"Wage Month",False,False),
                  ("TRRN",18,"TRRN",False,False),("ECR Id",14,"ECR Id",False,False),("LIN",14,"LIN",False,False),
                  ("EPF\nSubscribers",13,"Total Subscribers EPF",False,True),("EPS\nSubscribers",13,"Total Subscribers EPS",False,True),
                  ("EPF Total\nWages (Rs.)",16,"Total Wages EPF",True,False),("EPS Total\nWages (Rs.)",16,"Total Wages EPS",True,False),
                  ("Admin\nA/C.01",12,"Admin A/C.01",True,False),("Admin\nA/C.02",12,"Admin A/C.02",True,False),
                  ("Admin\nA/C.10",12,"Admin A/C.10",True,False),("Admin\nA/C.21",12,"Admin A/C.21",True,False),
                  ("Admin\nA/C.22",12,"Admin A/C.22",True,False),("Admin\nTotal (Rs.)",14,"Admin Total",True,False),
                  ("Employer\nA/C.01",12,"Employer A/C.01",True,False),("Employer\nA/C.02",12,"Employer A/C.02",True,False),
                  ("Employer\nA/C.10",12,"Employer A/C.10",True,False),("Employer\nA/C.21",12,"Employer A/C.21",True,False),
                  ("Employer\nA/C.22",12,"Employer A/C.22",True,False),("Employer\nTotal (Rs.)",15,"Employer Total",True,False),
                  ("Employee\nA/C.01",12,"Employee A/C.01",True,False),("Employee\nA/C.02",12,"Employee A/C.02",True,False),
                  ("Employee\nA/C.10",12,"Employee A/C.10",True,False),("Employee\nA/C.21",12,"Employee A/C.21",True,False),
                  ("Employee\nA/C.22",12,"Employee A/C.22",True,False),("Employee\nTotal (Rs.)",15,"Employee Total",True,False),
                  ("Grand\nTotal (Rs.)",16,"Grand Total",True,False),("Grand Total (In Words)",45,"Grand Total (Words)",False,False),
                  ("Total Remittance\nby Employer (Rs.)",18,"Total Remittance by Employer",True,False),
                  ("Total ECR\nAmount (Rs.)",16,"Total ECR Amount",True,False)]
            SECS=[("Establishment Information",1,9,"1F4E79"),("Subscribers & Wages",10,13,"1F4E79"),
                  ("Administration Charges",14,19,"375623"),("Employer's Share",20,25,"843C0C"),
                  ("Employee's Share",26,31,"7030A0"),("Totals",32,35,"1F4E79")]
            SCMAP={c:col for _,s,e,col in SECS for c in range(s,e+1)}
            ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(COLS))
            c=ws.cell(row=1,column=1,value="EMPLOYEES' PROVIDENT FUND — CHALLAN CONSOLIDATED REPORT")
            c.font=Font(name="Arial",bold=True,size=13,color="FFFFFF"); c.fill=PatternFill("solid",start_color="1F4E79")
            c.alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[1].height=28; ws.row_dimensions[2].height=16
            for name,s,e,color in SECS:
                ws.merge_cells(start_row=2,start_column=s,end_row=2,end_column=e); c=ws.cell(row=2,column=s,value=name)
                c.font=Font(name="Arial",bold=True,size=9,color="FFFFFF"); c.fill=PatternFill("solid",start_color=color); c.alignment=ctr; c.border=bdr
            ws.row_dimensions[3].height=38
            for ci,(hdr,width,_,_,_) in enumerate(COLS,1):
                ws.column_dimensions[get_column_letter(ci)].width=width; c=ws.cell(row=3,column=ci,value=hdr)
                c.font=Font(name="Arial",bold=True,size=9,color="FFFFFF"); c.fill=PatternFill("solid",start_color=SCMAP.get(ci,"1F4E79")); c.alignment=ctr; c.border=bdr
            for ri,rec in enumerate(records,1):
                er=ri+3; ws.row_dimensions[er].height=16; bg=["FFFFFF","EBF3FB"][ri%2]
                for ci,(_,_,key,ic,ii) in enumerate(COLS,1):
                    v=ri if key=="S.No" else rec.get(key,""); fmt=INR if ic else (INT if ii else None)
                    mc(er,ci,v,fg="000000",bg=bg,align=rgt if (ic or ii) else lft,fmt=fmt)
            tr=len(records)+4; ws.row_dimensions[tr].height=18
            ws.merge_cells(start_row=tr,start_column=1,end_row=tr,end_column=8); c=ws.cell(row=tr,column=1,value="GRAND TOTAL")
            c.font=Font(name="Arial",bold=True,size=10,color="1F4E79"); c.fill=PatternFill("solid",start_color="FFF2CC"); c.alignment=ctr; c.border=bdr
            for ci,(_,_,_,ic,ii) in enumerate(COLS,1):
                if ci not in (set(range(10,32))|{32,34,35}): continue
                cl=get_column_letter(ci); fmt=INR if ic else (INT if ii else None)
                c=ws.cell(row=tr,column=ci,value=f"=SUM({cl}4:{cl}{tr-1})")
                c.font=Font(name="Arial",bold=True,size=9); c.fill=PatternFill("solid",start_color="FFF2CC"); c.alignment=rgt; c.border=bdr
                if fmt: c.number_format=fmt
            ws.freeze_panes="A4"; buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

        st.title("📋 EPF Challan Consolidator")
        st.markdown("Upload any number of **EPF Combined Challan PDFs** to merge into a single formatted Excel report.")
        st.markdown("---")
        uploaded=st.file_uploader("Upload EPF Challan PDFs",type=["pdf"],accept_multiple_files=True,key="epf_uploader")
        if uploaded:
            valid,results=[],[]
            for f in uploaded:
                data,err=epf_extract(f)
                if err: results.append({"file":f.name,"ok":False,"detail":err})
                else:
                    data["Source File"]=f.name; valid.append(data); gt=data.get("Grand Total","")
                    results.append({"file":f.name,"ok":True,"detail":f"Estab: **{data.get('Establishment Code','')}** | Month: **{data.get('Wage Month','')}** | Grand Total: **{'₹'+f'{gt:,}' if isinstance(gt,int) else str(gt)}**"})
            st.markdown("### 📂 File Processing Results")
            for r in results:
                c1,c2,c3=st.columns([3,2,6]); c1.write(f"`{r['file']}`"); c2.write(f"{'✅' if r['ok'] else '❌'} {'Valid' if r['ok'] else 'Wrong Format'}"); c3.markdown(r["detail"])
            st.markdown("---")
            if valid:
                st.success(f"✅ **{len(valid)} valid challan(s)** ready.")
                if st.button("📥 Generate Consolidated Excel",type="primary",use_container_width=True,key="epf_gen"):
                    with st.spinner("Building..."): buf=epf_excel(valid)
                    st.download_button("⬇️ Download EPF_Challan_Consolidated.xlsx",data=buf,file_name="EPF_Challan_Consolidated.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)
                    k1,k2,k3,k4,k5=st.columns(5)
                    k1.metric("Establishments",len(valid)); k2.metric("Total EPF Members",f"{sum(r.get('Total Subscribers EPF',0) or 0 for r in valid):,}")
                    k3.metric("Total Wages",f"Rs.{sum(r.get('Total Wages EPF',0) or 0 for r in valid):,.0f}")
                    k4.metric("Employee Share",f"Rs.{sum(r.get('Employee Total',0) or 0 for r in valid):,.0f}")
                    k5.metric("Grand Total",f"Rs.{sum(r.get('Grand Total',0) or 0 for r in valid):,.0f}")
                    for i,r in enumerate(valid,1):
                        gt=r.get("Grand Total",0) or 0
                        with st.expander(f"{i}. {r.get('Establishment Code','')} — {r.get('Company Name','')} | {r.get('Wage Month','')} | Rs.{gt:,}"):
                            ca,cb,cc=st.columns(3)
                            ca.markdown(f"**Address:** {r.get('Address','')}"); ca.markdown(f"**TRRN:** `{r.get('TRRN','')}`"); ca.markdown(f"**LIN:** `{r.get('LIN','')}`")
                            cb.metric("EPF Subscribers",r.get("Total Subscribers EPF","")); cb.metric("EPS Subscribers",r.get("Total Subscribers EPS","")); cb.metric("Wages",f"Rs.{r.get('Total Wages EPF',0):,}")
                            cc.metric("Admin",f"Rs.{r.get('Admin Total',0):,}"); cc.metric("Employer",f"Rs.{r.get('Employer Total',0):,}"); cc.metric("Employee",f"Rs.{r.get('Employee Total',0):,}")
            else: st.error("No valid EPF challans found.")
        else:
            st.info("👆 Upload one or more EPF Challan PDFs above.")
            with st.expander("ℹ️ How it works"):
                st.markdown("1. **Upload** EPF Combined Challan PDFs\n2. **Parse** — each is validated automatically\n3. **Export** — consolidated Excel with one click\n4. Non-EPF PDFs are flagged and skipped")

    # ═══════════════════════════════════════════════════════════════════════════
    # ESIC
    # ═══════════════════════════════════════════════════════════════════════════
    elif active["id"] == "esic":
        import pdfplumber, pandas as pd, re
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        st.markdown(f'<h1 style="font-family:Syne,sans-serif;font-size:1.7rem;font-weight:800;color:{T["title_col"]};border-bottom:3px solid #c0392b;padding-bottom:.4rem;margin-bottom:.4rem;">🏥 ESIC Challan Extractor</h1>', unsafe_allow_html=True)
        st.markdown(f'<p style="color:{T["sub_col"]};margin-bottom:1rem;">Upload one or more ESIC Challan PDFs to extract and export data to Excel.</p>', unsafe_allow_html=True)

        EF={"Employer's Code No":"Employer Code No","Employer's Name":"Employer Name","Challan Period":"Challan Period",
            "Challan Number":"Challan Number","Challan Created Date":"Challan Created Date",
            "Challan Submitted Date":"Challan Submitted Date","Amount Paid":"Amount Paid",
            "Transaction Number":"Transaction Number","Transaction status":"Transaction Status"}

        def esic_extract(file_bytes):
            data={}
            with pdfplumber.open(BytesIO(file_bytes)) as pdf:
                text="\n".join(p.extract_text() or "" for p in pdf.pages)
            for field,col in EF.items():
                m=re.search(re.escape(field)+r"[\s:]*([^\n]+)",text,re.IGNORECASE)
                data[col]=m.group(1).strip().rstrip("*").strip() if m else ""
            return data

        def esic_excel(records):
            wb=Workbook(); ws=wb.active; ws.title="ESIC Challans"
            hdrs=["Source File"]+list(EF.values())
            HFn=Font(name="Arial",bold=True,color="FFFFFF",size=11)
            HF=PatternFill("solid",start_color="1A1A2E")
            HA=Alignment(horizontal="center",vertical="center",wrap_text=True)
            DF=Font(name="Arial",size=10); AF=PatternFill("solid",start_color="F2F2EF")
            CA=Alignment(horizontal="center",vertical="center")
            thin=Side(style="thin",color="CCCCCC"); B=Border(left=thin,right=thin,top=thin,bottom=thin)
            for ci,h in enumerate(hdrs,1):
                c=ws.cell(row=1,column=ci,value=h); c.font=HFn; c.fill=HF; c.alignment=HA; c.border=B
            ws.row_dimensions[1].height=30
            for ri,rec in enumerate(records,2):
                fill=PatternFill("solid",start_color="FFFFFF") if ri%2==0 else AF
                for ci,h in enumerate(hdrs,1):
                    c=ws.cell(row=ri,column=ci,value=rec.get(h,"")); c.font=DF; c.fill=fill; c.alignment=CA; c.border=B
                ws.row_dimensions[ri].height=20
            for i,w in enumerate([30,22,28,14,22,22,22,14,22,26],1): ws.column_dimensions[get_column_letter(i)].width=w
            tr=len(records)+2
            ws.cell(row=tr,column=1,value="TOTAL").fill=PatternFill("solid",start_color="C0392B")
            ws.cell(row=tr,column=1).font=Font(name="Arial",bold=True,color="FFFFFF"); ws.cell(row=tr,column=1).alignment=CA
            ac=list(EF.values()).index("Amount Paid")+2; gl=get_column_letter(ac)
            tc=ws.cell(row=tr,column=ac,value=f"=SUM({gl}2:{gl}{tr-1})")
            tc.font=Font(name="Arial",bold=True,color="FFFFFF"); tc.fill=PatternFill("solid",start_color="C0392B"); tc.alignment=CA; tc.border=B
            ws.freeze_panes="A2"; out=BytesIO(); wb.save(out); out.seek(0); return out

        uf=st.file_uploader("Upload ESIC Challan PDFs",type=["pdf"],accept_multiple_files=True,key="esic_uploader")
        if uf:
            st.markdown(f"**{len(uf)} file(s) uploaded**")
            recs=[]; errs=[]
            for f in uf:
                try: rec=esic_extract(f.read()); rec["_filename"]=f.name; recs.append(rec)
                except Exception as e: errs.append(f"{f.name}: {e}")
            for err in errs: st.error(f"⚠️ {err}")
            if recs:
                disp=[{"Source File":r["_filename"],**{k:v for k,v in r.items() if k!="_filename"}} for r in recs]
                st.markdown("### Preview")
                st.dataframe(pd.DataFrame(disp),use_container_width=True)
                try: ta=sum(float(r.get('Amount Paid',0) or 0) for r in disp); st.markdown(f"**{len(recs)} record(s)** | Total: ₹{ta:,.2f}")
                except: st.markdown(f"**{len(recs)} record(s) extracted**")
                if st.button("⬇ Download Excel",key="esic_dl"):
                    st.download_button("📥 Save Excel File",data=esic_excel(disp),file_name="ESIC_Challans.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.info("Upload PDFs above to get started.")

    # ═══════════════════════════════════════════════════════════════════════════
    # EXCEL CONSOLIDATOR
    # ═══════════════════════════════════════════════════════════════════════════
    elif active["id"] == "excel":
        import pandas as pd, numpy as np, warnings
        from io import BytesIO
        from pathlib import Path
        from bs4 import BeautifulSoup
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import xlrd
        warnings.filterwarnings("ignore")

        st.markdown(f"""<style>
        .stat-card{{background:{T['card_bg']};border-radius:12px;padding:1.1rem 1.4rem;border:1px solid {T['card_border']};border-left:4px solid #2d6a9f;margin-bottom:.5rem;}}
        .stat-card .label{{font-size:.76rem;color:{T['sub_col']};font-weight:500;text-transform:uppercase;letter-spacing:.05em;}}
        .stat-card .value{{font-size:1.6rem;font-weight:700;color:{T['title_col']};margin-top:.1rem;}}
        .stat-card .sub{{font-size:.8rem;color:{T['sub_col']};}}
        .file-badge{{display:inline-block;padding:2px 10px;border-radius:20px;font-size:.75rem;font-weight:600;margin:2px;}}
        .badge-html{{background:#dbeafe;color:#1d4ed8;}}.badge-xls{{background:#dcfce7;color:#166534;}}
        .badge-xlsx{{background:#fef9c3;color:#854d0e;}}.badge-csv{{background:#fce7f3;color:#9d174d;}}
        .step-box{{background:{T['card_bg']};border-radius:10px;padding:1rem 1.2rem;margin-bottom:.8rem;border:1px solid {T['card_border']};}}
        .step-num{{display:inline-block;width:26px;height:26px;background:#2d6a9f;color:white;border-radius:50%;text-align:center;line-height:26px;font-size:.8rem;font-weight:700;margin-right:8px;}}
        .error-box{{background:#fef2f2;border:1.5px solid #fca5a5;border-radius:10px;padding:.8rem 1.2rem;margin:.5rem 0;}}
        </style>""", unsafe_allow_html=True)

        def xdet(fb,fn):
            ext=Path(fn).suffix.lower()
            if ext==".csv": return "csv"
            if ext in (".xlsx",".xlsm"): return "xlsx"
            if fb[:4]==b"\xd0\xcf\x11\xe0": return "xls_real"
            snip=fb[:200].decode("utf-8",errors="ignore").lower()
            return "xls_html" if ("<table" in snip or "<html" in snip) else "xls_real"

        def xhtml(fb,sn):
            soup=BeautifulSoup(fb.decode("utf-8",errors="ignore"),"html.parser"); t=soup.find("table")
            if not t: raise ValueError("No HTML table found.")
            hr=t.find("tr"); hdrs=[th.get_text(strip=True) for th in hr.find_all(["th","td"])]
            rows=[[td.get_text(separator=" ",strip=True) for td in tr.find_all("td")] for tr in t.find_all("tr")[1:]]
            rows=[r for r in rows if any(c.strip() for c in r)]
            if not rows: raise ValueError("No data rows.")
            nc=max(len(r) for r in rows); hdrs=(hdrs+[""]*nc)[:nc]
            return pd.DataFrame(rows,columns=hdrs),"HTML-as-XLS"

        def xxls(fb,sn,sc):
            s=pd.read_excel(BytesIO(fb),engine="xlrd",sheet_name=None,dtype=str)
            if not s: raise ValueError("No sheets.")
            df=s[sc] if (sc and sc in s) else list(s.values())[0]
            return df.dropna(how="all").reset_index(drop=True),"XLS (Binary)"

        def xxlsx(fb,sn,sc):
            s=pd.read_excel(BytesIO(fb),engine="openpyxl",sheet_name=None,dtype=str)
            if not s: raise ValueError("No sheets.")
            df=s[sc] if (sc and sc in s) else list(s.values())[0]
            return df.dropna(how="all").reset_index(drop=True),"XLSX"

        def xcsv(fb,sn,enc="utf-8"):
            for e in [enc,"latin-1","cp1252"]:
                try: return pd.read_csv(BytesIO(fb),dtype=str,encoding=e).dropna(how="all").reset_index(drop=True),"CSV"
                except: continue
            raise ValueError("Could not decode CSV.")

        def xsheets(fb,fn):
            ext=Path(fn).suffix.lower()
            try:
                if ext in (".xlsx",".xlsm"): return pd.ExcelFile(BytesIO(fb),engine="openpyxl").sheet_names
                elif ext==".xls" and xdet(fb,fn)=="xls_real": return pd.ExcelFile(BytesIO(fb),engine="xlrd").sheet_names
            except: pass
            return []

        def xparse(fb,fn,sc=None):
            fmt=xdet(fb,fn)
            if fmt=="csv": return xcsv(fb,fn)
            elif fmt=="xlsx": return xxlsx(fb,fn,sc)
            elif fmt=="xls_html": return xhtml(fb,fn)
            else: return xxls(fb,fn,sc)

        def xbuild(combined,inc_summary):
            wb=Workbook()
            HF=PatternFill("solid",start_color="1F4E79",end_color="1F4E79"); HFn=Font(name="Arial",bold=True,color="FFFFFF",size=10)
            DF=Font(name="Arial",size=9); AF=PatternFill("solid",start_color="EBF3FB",end_color="EBF3FB")
            SF=PatternFill("solid",start_color="FFF9C4",end_color="FFF9C4"); SFn=Font(name="Arial",size=9,color="7B4F00")
            CTR=Alignment(horizontal="center",vertical="center"); LFT=Alignment(horizontal="left",vertical="center")
            thin=Side(style="thin",color="CCCCCC"); BDR=Border(left=thin,right=thin,top=thin,bottom=thin)
            ws=wb.active; ws.title="Consolidated Data"; cols=list(combined.columns)
            for ci,col in enumerate(cols,1):
                c=ws.cell(row=1,column=ci,value=col); c.font=HFn; c.fill=HF; c.alignment=CTR; c.border=BDR
                if col=="Source File": c.fill=PatternFill("solid",start_color="B8860B",end_color="B8860B")
            for ri,rd in enumerate(combined.itertuples(index=False),2):
                alt=ri%2==0
                for ci,val in enumerate(rd,1):
                    v="" if (val is None or (isinstance(val,float) and np.isnan(val))) else str(val)
                    c=ws.cell(row=ri,column=ci,value=v); c.border=BDR; c.font=DF
                    if cols[ci-1]=="Source File": c.fill=SF; c.font=SFn; c.alignment=CTR
                    else: c.fill=AF if alt else PatternFill(); c.alignment=CTR if ci<=3 else LFT
            for ci,col in enumerate(cols,1):
                ml=max([len(str(col))]+[len(str(ws.cell(row=ri,column=ci).value or "")) for ri in range(2,min(len(combined)+2,200))])
                ws.column_dimensions[get_column_letter(ci)].width=min(ml+3,45)
            ws.row_dimensions[1].height=22; ws.freeze_panes="B2"
            if inc_summary and "Source File" in combined.columns:
                ws2=wb.create_sheet("File Summary")
                for ci,h in enumerate(["Source File","Rows","Columns"],1):
                    c=ws2.cell(row=1,column=ci,value=h); c.font=HFn; c.fill=HF; c.alignment=CTR; c.border=BDR
                groups=combined.groupby("Source File",sort=False)
                for ri,(src,grp) in enumerate(groups,2):
                    for ci,v in enumerate([src,len(grp),len(combined.columns)-1],1):
                        c=ws2.cell(row=ri,column=ci,value=v); c.font=DF; c.alignment=CTR; c.border=BDR
                        if ri%2==0: c.fill=AF
                tr=len(groups)+2
                for ci,v in enumerate(["TOTAL",len(combined),""],1):
                    c=ws2.cell(row=tr,column=ci,value=v); c.font=Font(name="Arial",bold=True,color="FFFFFF",size=10)
                    c.fill=HF; c.alignment=CTR; c.border=BDR
                for ci,w in enumerate([40,12,12],1): ws2.column_dimensions[get_column_letter(ci)].width=w
                ws2.row_dimensions[1].height=22
            buf=BytesIO(); wb.save(buf); return buf.getvalue()

        if "xl_pf" not in st.session_state:
            st.session_state.xl_pf={}; st.session_state.xl_ff={}; st.session_state.xl_fe={}; st.session_state.xl_cd=None

        with st.sidebar:
            st.markdown("## ⚙️ Options"); st.markdown("---")
            out_fn=st.text_input("Output filename",value="Consolidated_Output")
            inc_sum=st.checkbox("Add File Summary sheet",value=True)
            st.markdown("### 🧹 Column Handling")
            col_strat=st.selectbox("Mismatched columns",["Union (keep all columns)","Intersection (common columns only)"])
            add_rn=st.checkbox("Add row number column",value=False)
            st.markdown("---"); st.markdown("### 📋 Formats")
            for fmt,badge in [("`.xls` Binary","badge-xls"),("`.xls` HTML-as-XLS","badge-html"),("`.xlsx`/`.xlsm`","badge-xlsx"),("`.csv`","badge-csv")]:
                st.markdown(f'<span class="file-badge {badge}">{fmt}</span>',unsafe_allow_html=True)

        st.markdown("""<div style="background:linear-gradient(135deg,#1e3a5f 0%,#2d6a9f 50%,#1a8cff 100%);border-radius:16px;padding:2rem;margin-bottom:1.5rem;color:white;">
            <h1 style="font-size:2rem;font-weight:700;margin:0 0 .3rem;">📊 Excel File Consolidator</h1>
            <p style="font-size:.95rem;opacity:.9;margin:0;">Merge multiple Excel / CSV files into one clean, formatted spreadsheet.</p>
        </div>""", unsafe_allow_html=True)

        st.markdown('<div class="step-box"><span class="step-num">1</span> <b>Upload your files</b></div>', unsafe_allow_html=True)
        uploaded=st.file_uploader("Drop files here",type=["xls","xlsx","xlsm","csv"],accept_multiple_files=True,label_visibility="collapsed",key="xl_uploader")

        if uploaded:
            st.markdown(f"**{len(uploaded)} file(s) selected**"); ss={}
            for uf in uploaded:
                fb=uf.read(); uf.seek(0); sheets=xsheets(fb,uf.name)
                if len(sheets)>1:
                    with st.expander(f"📋 `{uf.name}` — choose sheet"):
                        ss[uf.name]=st.selectbox(f"Sheet",sheets,key=f"xl_s_{uf.name}",label_visibility="collapsed")

            st.markdown('<div class="step-box"><span class="step-num">2</span> <b>Parse & preview</b></div>', unsafe_allow_html=True)
            if st.button("🔍 Parse All Files",key="xl_parse"):
                st.session_state.xl_pf={}; st.session_state.xl_ff={}; st.session_state.xl_fe={}; st.session_state.xl_cd=None
                prog=st.progress(0); stat=st.empty()
                for i,uf in enumerate(uploaded):
                    stat.text(f"Parsing {uf.name}…")
                    try:
                        fb=uf.read(); df,fmt=xparse(fb,uf.name,ss.get(uf.name))
                        st.session_state.xl_pf[uf.name]=df.dropna(how="all").reset_index(drop=True)
                        st.session_state.xl_ff[uf.name]=fmt
                    except Exception as e: st.session_state.xl_fe[uf.name]=str(e)
                    prog.progress((i+1)/len(uploaded))
                stat.empty(); prog.empty()

            for fn,df in st.session_state.xl_pf.items():
                fmt=st.session_state.xl_ff.get(fn,"")
                bc={"HTML-as-XLS":"badge-html","XLS (Binary)":"badge-xls","XLSX":"badge-xlsx","CSV":"badge-csv"}.get(fmt,"badge-xls")
                with st.expander(f"✅ `{fn}` — {len(df):,} rows × {len(df.columns)} cols"):
                    st.markdown(f'<span class="file-badge {bc}">{fmt}</span>',unsafe_allow_html=True)
                    st.dataframe(df.head(5),use_container_width=True,hide_index=True)

            for fn,err in st.session_state.xl_fe.items():
                st.markdown(f'<div class="error-box">❌ <b>{fn}</b>: {err}</div>',unsafe_allow_html=True)

            if st.session_state.xl_pf:
                st.markdown('<div class="step-box"><span class="step-num">3</span> <b>Consolidate</b></div>', unsafe_allow_html=True)
                if st.button("⚡ Consolidate All Files",key="xl_con"):
                    dfs=[]
                    for fn,df in st.session_state.xl_pf.items(): d=df.copy(); d.insert(0,"Source File",fn); dfs.append(d)
                    if col_strat.startswith("Union"): combined=pd.concat(dfs,ignore_index=True,sort=False)
                    else:
                        common=set(dfs[0].columns)
                        for d in dfs[1:]: common&=set(d.columns)
                        common=sorted(common,key=lambda c:list(dfs[0].columns).index(c) if c in dfs[0].columns else 999)
                        combined=pd.concat([d[list(common)] for d in dfs],ignore_index=True,sort=False)
                    if add_rn: combined.insert(1,"Row No.",range(1,len(combined)+1))
                    st.session_state.xl_cd=combined

                if st.session_state.xl_cd is not None:
                    combined=st.session_state.xl_cd
                    fc=combined["Source File"].nunique() if "Source File" in combined.columns else len(st.session_state.xl_pf)
                    c1,c2,c3,c4=st.columns(4)
                    for co,lbl,val,sub in [(c1,"Total Rows",f"{len(combined):,}","rows"),(c2,"Columns",f"{len(combined.columns):,}","cols"),(c3,"Files",str(fc),"merged"),(c4,"Cells",f"~{len(combined)*len(combined.columns)//1000}K","total")]:
                        co.markdown(f'<div class="stat-card"><div class="label">{lbl}</div><div class="value">{val}</div><div class="sub">{sub}</div></div>',unsafe_allow_html=True)
                    st.markdown("**Preview — first 50 rows**")
                    st.dataframe(combined.head(50),use_container_width=True,hide_index=True)
                    st.markdown('<div class="step-box"><span class="step-num">4</span> <b>Download</b></div>', unsafe_allow_html=True)
                    d1,d2=st.columns(2)
                    with d1:
                        with st.spinner("Building Excel..."): xlsx=xbuild(combined,inc_sum)
                        st.download_button("⬇️ Download as Excel",data=xlsx,file_name=f"{out_fn}.xlsx",
                                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    with d2:
                        st.download_button("⬇️ Download as CSV",data=combined.to_csv(index=False).encode("utf-8-sig"),
                                           file_name=f"{out_fn}.csv",mime="text/csv")
        else:
            st.markdown(f"""<div style="text-align:center;padding:3rem 1rem;">
                <div style="font-size:3.5rem;margin-bottom:1rem;">📂</div>
                <div style="font-size:1.05rem;font-weight:600;color:{T['title_col']};">Upload your Excel or CSV files above to get started</div>
                <div style="font-size:.88rem;margin-top:.4rem;color:{T['sub_col']};">Supports XLS (binary & HTML), XLSX, XLSM, and CSV</div>
            </div>""", unsafe_allow_html=True)
